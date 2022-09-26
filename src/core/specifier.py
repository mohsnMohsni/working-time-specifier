import json
from typing import Generator
from random import randint, shuffle

from jdatetime import date as j_date, set_locale
from openpyxl import load_workbook

from _calendar.manager import WorkingsDate


set_locale('fa_IR')


def generate_random_specifie_time(branches_col: int = 10, sum_of_random: int = 8) -> list:
    '''
    Generate list of randoms number that sum of them is specified in range.
    '''

    num_sum, numbers = int(), list()
    for index in range(branches_col):
        n = randint(0, sum_of_random - num_sum)
        numbers.append(n)
        num_sum += n
        if num_sum == sum_of_random:
            numbers += [0] * (branches_col - (index + 1))
            break
    shuffle(numbers)
    return numbers


def work_branch_handler(conf: dict, sum_of_random: int = 10) -> list:
    '''
    Read from conf to specify generated times should init into which branch of work.
    '''

    _times = generate_random_specifie_time(sum(conf.values()) * 2, sum_of_random=sum_of_random)
    return {
        6: _times.pop() if conf['mainSite'] else 0,
        7: _times.pop() if conf['mainSite'] else 0,
        8: _times.pop() if conf['mag'] else 0,
        9: _times.pop() if conf['mag'] else 0,
        10: _times.pop() if conf['managementPanel'] else 0,
        11: _times.pop() if conf['managementPanel'] else 0,
        12: _times.pop() if conf['workmanPanel'] else 0,
        13: _times.pop() if conf['workmanPanel'] else 0,
        14: _times.pop() if conf['clientPanel'] else 0,
        15: _times.pop() if conf['clientPanel'] else 0,
    }


class XlsxHandler:
    '''
    Open xls, retrive one left to last row and init data dpeneds on it then insert row into xls.
    '''

    def __init__(self) -> None:
        with open('config.json', 'r') as conf_file:
            conf = json.load(conf_file)
            self.xlsPath = conf.pop('xlsPath')
            self.conf = conf
        self.wb_obj = load_workbook(self.xlsPath)
        self.sheet_obj = self.wb_obj.active
        self.r_data = self.retrive_last_row_data()

    def retrive_last_row_data(self) -> dict:
        one_left_to_last_row = self.sheet_obj[self.sheet_obj.max_row - 1]
        return {i: cell.value for i, cell in enumerate(one_left_to_last_row)}

    def retrive_rows_to_be_filled(self) -> set:
        date_inserted_before = j_date(*map(int, self.r_data.get(5).split('.')))
        return WorkingsDate.retrive_working_dates(date_inserted_before.togregorian())

    def generate_rows_data(self) -> Generator:
        rows_to_be_filled = list(self.retrive_rows_to_be_filled())
        rows_to_be_filled.sort()
        index = int(self.r_data.get(0))
        for d in rows_to_be_filled:
            r_jalalit_date = j_date.fromtimestamp(int(d.strftime('%s')))
            index += 1
            sum_of_random = 8 if r_jalalit_date.jweekday() != 5 else 5
            yield {
                **self.r_data,
                0: index,
                4: r_jalalit_date.jweekday(),
                5: r_jalalit_date.strftime("%Y.%m.%d"),
                **work_branch_handler(self.conf, sum_of_random=sum_of_random),
            }

    def init_xls(self) -> None:
        data = self.generate_rows_data()
        index = self.sheet_obj.max_row
        for i, row in enumerate(data):
            self.sheet_obj.insert_rows(index + i)
            for j, cell_value in enumerate(row.values(), 1):
                ws_cell = self.sheet_obj.cell(row=index + i, column=j)
                ws_cell.value = cell_value
        self.wb_obj.save(self.xlsPath)
